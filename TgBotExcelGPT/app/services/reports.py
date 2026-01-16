import json
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook

from app.services import google_drive


def _project_reports_dir(project_id: str) -> Path:
    base = google_drive.LOCAL_STORAGE_ROOT / project_id / "reports"
    base.mkdir(parents=True, exist_ok=True)
    return base


def save_procurement_report(project_id: str, items: List[Dict]) -> Dict:
    reports_dir = _project_reports_dir(project_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость закупки"
    headers = ["Материал", "Характеристики", "Кол-во", "Ед. изм.", "Примечание"]
    ws.append(headers)
    for item in items:
        ws.append(
            [
                item.get("material") or item.get("name"),
                item.get("characteristics"),
                item.get("quantity"),
                item.get("unit"),
                item.get("note"),
            ]
        )
    excel_path = reports_dir / f"procurement_{project_id}.xlsx"
    wb.save(excel_path)

    json_path = reports_dir / f"procurement_{project_id}.json"
    json_path.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"excel": str(excel_path), "json": str(json_path)}


def save_discrepancy_report(project_id: str, rows: List[Dict]) -> Dict:
    reports_dir = _project_reports_dir(project_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Расхождения"
    headers = ["Позиция", "В спецификации", "По чертежам", "Отклонение", "Ед. изм."]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row.get("position"),
                row.get("spec"),
                row.get("drawings"),
                row.get("delta"),
                row.get("unit"),
            ]
        )
    excel_path = reports_dir / f"discrepancies_{project_id}.xlsx"
    wb.save(excel_path)

    json_path = reports_dir / f"discrepancies_{project_id}.json"
    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"excel": str(excel_path), "json": str(json_path)}


def save_supplier_summary(project_id: str, rows: List[Dict]) -> Dict:
    reports_dir = _project_reports_dir(project_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Поставщики"
    headers = [
        "Материал",
        "Кол-во",
        "Ед. изм.",
        "Поставщик",
        "Цена за ед.",
        "Итого",
        "Срок (дни)",
        "Канал",
        "Статус",
        "Лучшая цена",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row.get("material"),
                row.get("quantity"),
                row.get("unit"),
                row.get("supplier"),
                row.get("unit_price"),
                row.get("total"),
                row.get("lead_time_days"),
                row.get("channel"),
                row.get("status"),
                "да" if row.get("best") else "",
            ]
        )
    excel_path = reports_dir / f"suppliers_{project_id}.xlsx"
    wb.save(excel_path)

    json_path = reports_dir / f"suppliers_{project_id}.json"
    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"excel": str(excel_path), "json": str(json_path)}


def save_final_overview(
    project_id: str,
    discrepancies: List[Dict],
    procurement: List[Dict],
    supplier_rows: List[Dict],
) -> Dict:
    reports_dir = _project_reports_dir(project_id)
    wb = Workbook()

    ws_diff = wb.active
    ws_diff.title = "Расхождения"
    ws_diff.append(["Позиция", "В спецификации", "По чертежам", "Отклонение", "Ед. изм."])
    for row in discrepancies:
        ws_diff.append(
            [
                row.get("position"),
                row.get("spec"),
                row.get("drawings"),
                row.get("delta"),
                row.get("unit"),
            ]
        )

    ws_proc = wb.create_sheet("Ведомость закупки")
    ws_proc.append(["Материал", "Характеристики", "Кол-во", "Ед. изм.", "Примечание"])
    for item in procurement:
        ws_proc.append(
            [
                item.get("material") or item.get("name"),
                item.get("characteristics"),
                item.get("quantity"),
                item.get("unit"),
                item.get("note"),
            ]
        )

    ws_sup = wb.create_sheet("Поставщики")
    ws_sup.append(
        [
            "Материал",
            "Кол-во",
            "Ед. изм.",
            "Поставщик",
            "Цена за ед.",
            "Итого",
            "Срок (дни)",
            "Канал",
            "Статус",
            "Лучшая цена",
        ]
    )
    for row in supplier_rows:
        ws_sup.append(
            [
                row.get("material"),
                row.get("quantity"),
                row.get("unit"),
                row.get("supplier"),
                row.get("unit_price"),
                row.get("total"),
                row.get("lead_time_days"),
                row.get("channel"),
                row.get("status"),
                "да" if row.get("best") else "",
            ]
        )

    excel_path = reports_dir / f"final_{project_id}.xlsx"
    wb.save(excel_path)
    payload = {
        "discrepancies": discrepancies,
        "procurement": procurement,
        "suppliers": supplier_rows,
    }
    json_path = reports_dir / f"final_{project_id}.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"excel": str(excel_path), "json": str(json_path)}

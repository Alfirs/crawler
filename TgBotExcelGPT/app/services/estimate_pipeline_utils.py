import re
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


def _map_header(value: Optional[str], index: int) -> str:
    text = (str(value).strip().lower() if value is not None else "").replace(".", "")
    if "наимен" in text or "name" in text:
        return "name"
    if "кол" in text:
        return "quantity"
    if "ед" in text or "unit" in text:
        return "unit"
    if "прим" in text or "note" in text:
        return "note"
    if "хар" in text:
        return "characteristics"
    return f"col_{index}"


def _coerce_float(value: Optional[str]) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        digits = re.findall(r"\d+(?:\.\d+)?", text)
        if digits:
            try:
                return float(digits[0])
            except ValueError:
                return None
        return None


def load_estimate_rows(file_link: str) -> List[Dict]:
    path = Path(file_link)
    if not path.exists():
        return []
    wb = load_workbook(filename=str(path), data_only=True)
    rows: List[Dict] = []
    for sheet in wb.worksheets:
        header_map: List[str] = []
        header_row_index: Optional[int] = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if header_row_index is None:
                if not any(row):
                    continue
                header_row_index = idx
                header_map = [_map_header(cell, i) for i, cell in enumerate(row)]
                continue
            if not any(row):
                continue
            entry: Dict[str, Optional[str]] = {
                "row_number": idx,
                "sheet": sheet.title,
            }
            for col_index, key in enumerate(header_map):
                value = row[col_index] if col_index < len(row) else None
                entry[key] = value
            entry.setdefault("name", "")
            entry.setdefault("quantity", None)
            entry.setdefault("unit", "")
            entry.setdefault("note", "")
            entry.setdefault("characteristics", "")
            rows.append(entry)
        if rows:
            break
    return rows


def scan_estimate_issues(file_link: str) -> List[Dict]:
    rows = load_estimate_rows(file_link)
    issues: List[Dict] = []
    for row in rows:
        row_number = row.get("row_number", "?")
        name = str(row.get("name") or "").strip()
        if not name:
            issues.append({"row": row_number, "problem": "Пустое наименование"})
        quantity = row.get("quantity")
        quantity_value = _coerce_float(quantity)
        if quantity_value is None:
            issues.append({"row": row_number, "problem": "Не указано количество"})
        unit = str(row.get("unit") or "").strip()
        if not unit:
            issues.append({"row": row_number, "problem": "Не указана единица измерения"})
    return issues


SKIP_KEYWORDS = ["монтаж", "демонтаж", "работ", "коэфф", "наклад", "индекс", "услуг"]


def is_material_row(row: Dict) -> bool:
    name = str(row.get("name") or "").lower()
    if not name:
        return False
    return not any(keyword in name for keyword in SKIP_KEYWORDS)


def normalize_name(name: str) -> str:
    text = re.sub(r"[^\w\s/]", " ", name.lower())
    text = text.replace("пнд", "пнд")
    tokens = sorted(set(filter(None, text.split())))
    return " ".join(tokens)


def generate_procurement_list(file_link: str) -> List[Dict]:
    rows = load_estimate_rows(file_link)
    aggregated: Dict[str, Dict] = {}
    for row in rows:
        if not is_material_row(row):
            continue
        qty = _coerce_float(row.get("quantity"))
        if qty is None:
            continue
        unit = (str(row.get("unit")) if row.get("unit") is not None else "").strip()
        characteristics = (str(row.get("characteristics")) if row.get("characteristics") else "").strip()
        note = (str(row.get("note")) if row.get("note") else "").strip()
        normalized = normalize_name(str(row.get("name") or ""))
        key = "|".join([normalized, unit, characteristics])
        record = aggregated.setdefault(
            key,
            {
                "material": str(row.get("name") or "").strip(),
                "characteristics": characteristics,
                "quantity": 0.0,
                "unit": unit,
                "note": note,
            },
        )
        record["quantity"] += qty
        if not record["characteristics"] and characteristics:
            record["characteristics"] = characteristics
        if not record["note"] and note:
            record["note"] = note
    # Round quantities for readability
    for record in aggregated.values():
        record["quantity"] = round(record["quantity"], 3)
    return list(aggregated.values())

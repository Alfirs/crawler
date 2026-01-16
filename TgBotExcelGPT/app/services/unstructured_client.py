import asyncio
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from pypdf import PdfReader


def _parse_pdf(path: Path) -> Dict:
    reader = PdfReader(str(path))
    text_parts: List[str] = []
    for page in reader.pages:
        try:
            text_parts.append(page.extract_text() or "")
        except Exception:  # pragma: no cover - defensive
            continue
    return {"file_link": str(path), "text": "\n".join(text_parts), "tables": []}


def _parse_excel(path: Path) -> Dict:
    wb = load_workbook(filename=str(path), data_only=True)
    tables: List[List[str]] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if not any(row):
                continue
            tables.append([str(cell) if cell is not None else "" for cell in row])
    text_rows = ["\t".join(row) for row in tables]
    return {"file_link": str(path), "text": "\n".join(text_rows), "tables": tables}


def _parse_plain(path: Path) -> Dict:
    return {"file_link": str(path), "text": path.read_text(encoding="utf-8", errors="ignore"), "tables": []}


async def parse_document(file_link: str) -> Dict:
    """
    Simplified replacement for Unstructured.io parsing.
    """
    path = Path(file_link)
    if not path.exists():
        return {"file_link": file_link, "text": "", "tables": [], "error": "file_not_found"}

    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return await asyncio.to_thread(_parse_pdf, path)
    if suffix in {".xlsx", ".xlsm"}:
        return await asyncio.to_thread(_parse_excel, path)
    return await asyncio.to_thread(_parse_plain, path)
